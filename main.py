"""
Task:
[v] xls转换为xlsx
[v] 输入列数字母
[v] 添加选择空表格的按钮
[v] 优化界面文字
[v] 异常处理
[ ] 美化界面
"""

from tkinter import *
from tkinter import filedialog
from tkinter import messagebox
import xlsTOxlsx
import sys
from openpyxl import *
import os.path
import re
import shutil

root = Tk()

root.title("GradesMover App")

file_place = ""
empty_file_place = ""
folder_place = ""


# def find_dir(fp):
#     result = ""
#     ptr = 0
#     for index in range(len(fp)):
#         if fp[index] == "\\"



def chose_file():
    global file_place
    global student_file_name
    file_name = filedialog.askopenfilename(filetypes=[("XLS", ".xls")])
    # xlsTOxlsx.reverse_file(file_name)
    # file_place = file_name + "x"
    # # print(file_name)
    xlsTOxlsx.reverse_file(file_name)
    file_place = file_name + "x"
    file_place = os.path.normpath(file_place)
    # 匹配文件名称
    match = re.search(r"[^\\]+$", file_place)
    student_file_name.set(match.group()[:-1])
    # print(file_place)

# def chose_folder():
#     global folder_place
#     file_dir = filedialog.askdirectory()
#     os.chdir(file_dir)
#     folder_place = file_dir
#     # print(file_dir)


def chose_empty_file():
    global empty_file_place
    global folder_place
    file_name = filedialog.askopenfilename(filetypes=[("XLS", ".xls")])
    xlsTOxlsx.reverse_file(file_name)
    empty_file_place = file_name + "x"
    folder_place = os.path.split(empty_file_place)[0]

    empty_file_place = os.path.normpath(empty_file_place)
    folder_place = os.path.normpath(folder_place)

    # 匹配文件名称
    empty_file_name.set(re.search(r"[^\\]+$", empty_file_place).group()[:-1])


empty_file_name = StringVar()
empty_file_name.set("未选择文件！")
student_file_name = StringVar()
student_file_name.set("未选择文件！")


# Label(root, text="选择数据来源文件夹：").grid(row=0)
# Button(root, text="选择文件夹", command=chose_folder).grid(row=0, column=2)
Label(root, text="选择带有成绩的表格：").grid(row=2, pady=6)
Label(root, textvariable=student_file_name).grid(row=2, column=2)
Button(root, text="选择文件", command=chose_file).grid(row=2, column=3, pady=6)
Label(root, text="选择空表：").grid(row=4, pady=10)
Label(root, textvariable=empty_file_name).grid(row=4, column=2)
Button(root, text="选择文件", command=chose_empty_file).grid(row=4, column=3, pady=6)

Label(root, text="输入带有成绩的表格里的的学号列\n（大写字母）：").grid(row=6, column=0, pady=6)
Label(root, text="输入带有成绩的表格的分数列\n（大写字母）：").grid(row=8, column=0, pady=6)
Label(root, text="输入空表的学号列\n（大写字母）：").grid(row=10, column=0, pady=6)
Label(root, text="输入空表的成绩列\n（大写字母）：").grid(row=12, column=0, pady=6)

Label(root).grid(row=2, column=4, padx=10)

student_col_char = StringVar()
student_col = IntVar()
score_col_char = StringVar()
score_col = IntVar()
empty_student_col_char = StringVar()
empty_student_col = IntVar()
empty_score_col_char = StringVar()
empty_score_col = IntVar()

# 默认值
empty_student_col_char.set("A")
empty_score_col_char.set("D")

entry1 = Entry(root, textvariable=student_col_char)
entry2 = Entry(root, textvariable=score_col_char)
entry3 = Entry(root, textvariable=empty_student_col_char)
entry4 = Entry(root, textvariable=empty_score_col_char)

entry1.grid(row=6, column=3, padx=10, pady=5)
entry2.grid(row=8, column=3, padx=10, pady=5)
entry3.grid(row=10, column=3, padx=10, pady=5)
entry4.grid(row=12, column=3, padx=10, pady=5)


def show():
    # print(f"学号：{student_col.get()}")
    # print(f"分数：{score_col.get()}")
    root.quit()


Button(root, text="确定", width=10, command=show) \
    .grid(row=14, column=0, sticky=W, padx=10, pady=5)

Button(root, text="取消", width=10, command=sys.exit) \
    .grid(row=14, column=3, sticky=E, padx=10, pady=5)

root.mainloop()

# ------------------------------------------------------------------------ #

try:

    student_col.set(ord(student_col_char.get()) - 65)
    score_col.set(ord(score_col_char.get()) - 65)
    empty_student_col.set(ord(empty_student_col_char.get()) - 65)
    empty_score_col.set(ord(empty_score_col_char.get()) - 65)

    # print("注：数据储存到fin.xlsx\n数据从00.xlsx读取")
    # stra = input("调用表名（不加后缀）：")
    # fro = int(input("学号列："))
    student_col_var = student_col.get()
    # to = int(input("得分："))
    score_col_var = score_col.get()
    # numb = float(input())

    empty_workbook = load_workbook(f"{empty_file_place}", data_only=True)
    score_workbook = load_workbook(file_place, data_only=True)

    table_empty = empty_workbook["Sheet1"]
    table_score = score_workbook["Sheet1"]

    score_dic = {}
    empty_student_list = []


    for student in range(1, table_score.max_row):
        score_dic[str(list(table_score.rows)[student][student_col_var].value)] = str(list(table_score.rows)[student][score_col_var].value)


    for student in range(1, table_empty.max_row):
        empty_student_list.append(str(list(table_empty.columns)[empty_student_col.get()][student].value))

    keys = list(score_dic.keys())
    # print(list(score_dic.values()))

    # keys: dic_a:  a表
    # list_a:       总表

    ind = -1

    score_more_list = []
    empty_more_list = []
    no_score_list = []

    # print("\n\n------------------\n------keys-------------:\n", keys)
    # print("\n\n----------score_dic--------------\n", score_dic)

    for student in keys:
        if student != '' and student != "None" and score_dic[student] != 'None' and score_dic[student] != '-' and score_dic[student] != '':
            try:
                ind = empty_student_list.index(student)
            except:
                score_more_list.append(student) # 成绩表多人
                continue
            # print(ind)

            table_empty.cell(row=ind + 2, column=empty_score_col.get()+1).value = ""
            table_empty.cell(row=ind + 2, column=empty_score_col.get()+1).value = round(float(score_dic[student]), 0)
            empty_student_list[ind] = "None"
            # / numb
        else:
            if student != "None":
                no_score_list.append(student)
                empty_student_list[ind] = "None"

    shutil.copy(empty_file_place, folder_place + "/output.xlsx")
    empty_workbook.save(f"{folder_place}/output.xlsx")
    os.remove(empty_file_place)
    os.remove(file_place)

    window = Tk()
    window.withdraw()  # 退出默认 tk 窗口
    
    showinfo_text = "成功！"
    # 成绩表多人
    if (len(score_more_list) > 0):
        showinfo_text += f"\n\n警告：学号为 {score_more_list} 的成绩无法录入到空白表格。"
    
    # 空白表多人
    for each in empty_student_list:
        if each != "None":
            empty_more_list.append(each)
    if (len(empty_more_list) > 0):
        showinfo_text += f"\n\n警告：学号为 {empty_more_list} 的人不存在。"

    # 成绩为空
    if (len(no_score_list) > 0):
        showinfo_text += f"\n\n警告：学号为 {no_score_list} 的人无成绩。"
    
    # print(empty_student_list)

    messagebox.showinfo('Successful!', showinfo_text)
    sys.exit()
# ------------------------------------------------------------------------ #

except Exception as reason:
    messagebox.showerror(title="运行时错误！", message=f"""运行时出现错误：
可能由以下原因导致：
1. 文件损坏；
2. 文件的工作表名称（也就是excel中切换不同表格的那个）不是“Sheet1”，需要手动更改；
3. 输入数据有误;
4. 代码出bug了等。


错误原因：{reason}


注：如果错误原因为”ord() expected a character,but string of length 0 found“，可能原因是文件无法访问。
""")

